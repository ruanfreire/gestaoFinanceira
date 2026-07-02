import json
from pathlib import Path
from collections import Counter

def read(p):
    return Path(p).read_text(encoding='utf-8', errors='replace')

# inicial.json
p = Path(r'c:\Users\ruanf\OneDrive\Área de Trabalho\GestaoFinanceira\inicial.json')
data = json.loads(p.read_text(encoding='utf-8'))
items = data['data']['empresa']['nf_lista']['items']
print('=== inicial.json ===')
print('total:', len(items))
print('status:', Counter(i.get('status_emissao') for i in items))
print('codigo_servico:', Counter(i.get('codigo_servico') for i in items))
print('sample cancelada:', next((i for i in items if 'CANCEL' in str(i.get('status_emissao','')).upper()), None))

# Asaas
asaas = read(r'c:\Users\ruanf\OneDrive\Área de Trabalho\Extrato Asaas (1).csv')
lines = [l for l in asaas.splitlines() if l.strip()]
print('\n=== Asaas ===', 'lines', len(lines))
tipos = Counter(); lanc = Counter()
for l in lines:
    if l.startswith('"01/') or l.startswith('"02/') or l.startswith('"0'):
        parts = l.split(',')
        if len(parts) > 11:
            tipos[parts[2].strip('"')] += 1
            lanc[parts[11].strip('"')] += 1
print('tipos:', tipos.most_common(8))
print('tipo_lancamento:', lanc.most_common(4))

# Nubank files
for name in [
    r'c:\Users\ruanf\OneDrive\Área de Trabalho\663d3443-1768-4fe4-abe6-37b27b39172f-2026-06-01-2026-07-01.csv',
    r'c:\Users\ruanf\OneDrive\Área de Trabalho\Nubank_2026-06-02.csv',
]:
    print('\n===', Path(name).name, '===')
    print(read(name)[:800])

# xlsx
try:
    import openpyxl
    xlsx = Path(r'c:\Users\ruanf\OneDrive\Área de Trabalho\Ana Luisa_Fluxo de Caixa_modelo.xlsx')
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    print('\n=== XLSX sheets ===', wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'--- {sn} {ws.max_row}x{ws.max_column} ---')
        for r in range(1, min(25, ws.max_row+1)):
            row = [ws.cell(r,c).value for c in range(1, min(10, ws.max_column+1))]
            if any(v is not None for v in row):
                print(r, row)
except Exception as e:
    print('xlsx error:', e)
